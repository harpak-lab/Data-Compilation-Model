import pandas as pd
from openpyxl import load_workbook

def get_green_highlighted_rows(filepath, sheet_name):
    """Extracts only the green-highlighted rows from an Excel sheet and returns them as a DataFrame."""
    wb = load_workbook(filename=filepath, data_only=True)
    ws = wb[sheet_name]

    green_rows = [] # store highlighted entries
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.fill.start_color.index == "FF92D050": # hex for green
                green_rows.append(cell.row)
    
    df = pd.read_excel(filepath, sheet_name=sheet_name)
    
    # filter only green-highlighted rows
    df_filtered = df.iloc[[r - 2 for r in green_rows]]
    
    return df_filtered

def save_green_highlighted_rows(input_filepath, sheet_name, output_filepath):
    df_filtered = get_green_highlighted_rows(input_filepath, sheet_name)
    df_filtered.to_excel(output_filepath, index=False)
    print(f"Green-highlighted rows saved to {output_filepath}")

def filter_my_spreadsheet(my_spreadsheet, filtered_spreadsheet, output_filepath):
    df_my = pd.read_excel(my_spreadsheet)
    df_filtered = pd.read_excel(filtered_spreadsheet)
    
    # safety check: "name" column exists
    if "Name" not in df_my.columns or "Name" not in df_filtered.columns:
        raise ValueError("Missing 'Name' column in one of the spreadsheets.")
    
    # spreadsheet should only include species in filtered spreadsheet
    df_my_filtered = df_my[df_my["Name"].isin(df_filtered["Name"])]
    
    df_my_filtered.to_excel(output_filepath, index=False)
    print(f"Filtered spreadsheet saved to {output_filepath}")

def compare_values(reference_spreadsheet, my_filtered_spreadsheet, output_filepath):
    df_ref = pd.read_excel(reference_spreadsheet)
    df_my = pd.read_excel(my_filtered_spreadsheet)
    
    # safety check: "name" column exists
    if "Name" not in df_ref.columns or "Name" not in df_my.columns:
        raise ValueError("Missing 'Name' column in one of the spreadsheets.")
    
    # comparison dataframe
    df_comparison = pd.DataFrame(columns=["Name"] + [col for col in df_ref.columns if col != "Name"])
    df_comparison["Name"] = df_my["Name"]

    # compare vals
    for col in df_ref.columns:
        if col != "Name" and col in df_my.columns:
            df_my[col] = df_my[col].astype(str).str.strip().str.lower()
            df_ref[col] = df_ref[col].astype(str).str.strip().str.lower()

            df_comparison[col] = df_my[col].where(df_my[col] == "-", df_my[col].where(df_my[col] == df_ref[col], df_ref[col].where(df_ref[col] == 'nan', "-").fillna("-").replace("nan", "-")))

    # Save the comparison results
    df_comparison.to_excel(output_filepath, index=False)
    print(f"Comparison results saved to {output_filepath}")

big_spreadsheet = "Froggy_Spreadsheet.xlsx"  # reference spreadsheet
output_spreadsheet = "filtered_big_spreadsheet.xlsx"
my_spreadsheet = "results.xlsx"
filtered_output = "filtered_results.xlsx"
comparison_output = "final_comparison.xlsx"

# full pipeline
save_green_highlighted_rows(big_spreadsheet, "All Frogs", output_spreadsheet)
filter_my_spreadsheet(my_spreadsheet, output_spreadsheet, filtered_output)
compare_values(output_spreadsheet, filtered_output, comparison_output)
