import pandas as pd
from openpyxl import load_workbook
import os
import numpy as np

def filter_big_spreadsheet(input_filepath, sheet_name, output_filepath):
    # gets just green highlighted rows and returns them as dataframe
    wb = load_workbook(filename=input_filepath, data_only=True)
    ws = wb[sheet_name]

    green_rows = [] # store highlighted entries
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.fill.start_color.index == "FF92D050": # hex for green
                green_rows.append(cell.row)
    
    df = pd.read_excel(input_filepath, sheet_name=sheet_name)
    
    # filter only green-highlighted rows
    df_filtered = df.iloc[[r - 2 for r in green_rows]]

    # uncertainty cols -- make blanks 0 for cross verification
    uncertainty_columns = ["+/- SVL Male (mm)", "+/- SVL Female (mm)", "+/- SVL Adult (mm)", "+/- Egg Diameter (mm)"]
    for col in uncertainty_columns:
        if col in df_filtered.columns:
            df_filtered.loc[:, col] = df_filtered[col].fillna(0)


    df_filtered.to_excel(output_filepath, index=False)

def filter_my_spreadsheet(my_spreadsheet, filtered_spreadsheet, output_filepath):
    df_my = pd.read_excel(my_spreadsheet)
    df_filtered = pd.read_excel(filtered_spreadsheet)
    
    # safety check: "name" column exists
    if "Name" not in df_my.columns or "Name" not in df_filtered.columns:
        raise ValueError("Missing 'Name' column in one of the spreadsheets.")
    
    # spreadsheet should only include species in filtered spreadsheet
    df_my_filtered = df_my[df_my["Name"].isin(df_filtered["Name"])]
    df_my_filtered.to_excel(output_filepath, index=False)

def compare_values(reference_spreadsheet, my_filtered_spreadsheet, output_filepath):
    df_ref = pd.read_excel(reference_spreadsheet)
    df_filtered = pd.read_excel(my_filtered_spreadsheet)

    # merge by name
    df_merged = pd.merge(df_ref, df_filtered, on="Name", suffixes=("_ref", "_my"), how="left")

    df_comparison = pd.DataFrame()
    df_comparison["Name"] = df_merged["Name"]

    groups = [
        {"columns": ['SVL Male (mm)', '+/- SVL Male (mm)'], 'type': 'value_uncertainty'},
        {"columns": ['SVL Female (mm)', '+/- SVL Female (mm)'], 'type': 'value_uncertainty'},
        {"columns": ['Avg SVL Adult (mm)', '+/- SVL Adult (mm)'], 'type': 'value_uncertainty'},
        {"columns": ['Avg Egg Diameter (mm)', '+/- Egg Diameter (mm)'], 'type': 'value_uncertainty'},
        {"columns": ['Min Egg Clutch', 'Max Egg Clutch'], 'type': 'min_max'},
    ]

    def is_valid(value):
        if pd.isna(value): return False
        if isinstance(value, str) and value.strip() == '-': return False
        try: float(value); return True
        except: return False

    def ranges_overlap(a_low, a_high, b_low, b_high):
        return a_low <= b_high and b_low <= a_high

    for group in groups:
        cols = group['columns']
        group_type = group['type']

        for i, row in df_merged.iterrows():
            if group_type == 'value_uncertainty':
                ref_val, ref_unc = row[f"{cols[0]}_ref"], row[f"{cols[1]}_ref"]
                my_val, my_unc = row[f"{cols[0]}_my"], row[f"{cols[1]}_my"]
            else:
                ref_val, ref_unc = row[f"{cols[0]}_ref"], row[f"{cols[1]}_ref"]
                my_val, my_unc = row[f"{cols[0]}_my"], row[f"{cols[1]}_my"]

            if not all(is_valid(v) for v in [ref_val, ref_unc, my_val, my_unc]):
                df_comparison.at[i, cols[0]] = "-"
                df_comparison.at[i, cols[1]] = "-"
                continue

            ref_val_f, ref_unc_f = float(ref_val), float(ref_unc)
            my_val_f, my_unc_f = float(my_val), float(my_unc)

            if ref_val_f == my_val_f and ref_unc_f == my_unc_f:
                df_comparison.at[i, cols[0]] = ref_val_f
                df_comparison.at[i, cols[1]] = ref_unc_f
            else:
                if group_type == 'value_uncertainty':
                    ref_low, ref_high = ref_val_f - ref_unc_f, ref_val_f + ref_unc_f
                    my_low, my_high = my_val_f - my_unc_f, my_val_f + my_unc_f
                else:
                    ref_low, ref_high = ref_val_f, ref_unc_f
                    my_low, my_high = my_val_f, my_unc_f

                if ranges_overlap(ref_low, ref_high, my_low, my_high):
                    df_comparison.at[i, cols[0]] = "overlap"
                    df_comparison.at[i, cols[1]] = "overlap"
                else:
                    df_comparison.at[i, cols[0]] = "invalid"
                    df_comparison.at[i, cols[1]] = "invalid"

    df_comparison.to_excel(output_filepath, index=False)
    print(f"Comparative data saved to {output_filepath}")

def compare_altitudes():
    analysis_df = pd.read_excel("results/froggy_analysis_results.xlsx")
    reference_df = pd.read_excel("data/Reference_Froggy_Spreadsheet.xlsx")
    cross_verification_df = pd.read_excel("results/cross_verification_results.xlsx")

    result = pd.DataFrame(columns=["Min Altitude", "Max Altitude"])

    def safe_float(val):
        try:
            return float(val)
        except (ValueError, TypeError):
            return np.nan

    for i in range(len(analysis_df)):
        a_min = safe_float(analysis_df.at[i, "Min Altitude"])
        a_max = safe_float(analysis_df.at[i, "Max Altitude"])
        r_min = safe_float(reference_df.at[i, "Min Altitude"])
        r_max = safe_float(reference_df.at[i, "Max Altitude"])

        if pd.isna(a_min) or pd.isna(a_max) or pd.isna(r_min) or pd.isna(r_max): # missing vals
            result.loc[i] = ["-", "-"]
        elif a_min == r_min and a_max == r_max: # exact match
            result.loc[i] = [a_min, a_max]
        elif max(a_min, r_min) <= min(a_max, r_max): # overlap
            result.loc[i] = ["overlap", "overlap"]
        else: # no overlap
            result.loc[i] = ["invalid", "invalid"]

    cross_verification_df["Min Altitude"] = result["Min Altitude"]
    cross_verification_df["Max Altitude"] = result["Max Altitude"]

    if os.path.exists("results/cross_verification_results"):
        os.remove("results/cross_verification_results")

    cross_verification_df.to_excel("results/cross_verification_results.xlsx", index=False)

big_spreadsheet = "data/Reference_Froggy_Spreadsheet.xlsx"  # reference spreadsheet
output_spreadsheet = "filtered_big_spreadsheet.xlsx"
my_spreadsheet = "results/froggy_analysis_results.xlsx"
filtered_output = "filtered_results.xlsx"
comparison_output = "results/cross_verification_results.xlsx"

# full pipelin
filter_big_spreadsheet(big_spreadsheet, "All Frogs", output_spreadsheet)
filter_my_spreadsheet(my_spreadsheet, output_spreadsheet, filtered_output)
compare_values(output_spreadsheet, filtered_output, comparison_output)

# get rid of intermediary spreadsheets
if os.path.exists(output_spreadsheet):
    os.remove(output_spreadsheet)
if os.path.exists(filtered_output):
    os.remove(filtered_output)

compare_altitudes()