import pandas as pd
from openpyxl import load_workbook

def get_green_highlighted_names(filepath, sheet_name="All Frogs"):
    wb = load_workbook(filename=filepath, data_only=True)
    ws = wb[sheet_name]

    green_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.fill.start_color.index == "FF92D050":  # hex for green
                green_rows.append(cell.row)

    df = pd.read_excel(filepath, sheet_name=sheet_name)
    df_filtered = df.iloc[[r - 2 for r in green_rows]]  # adjust for 0-based indexing

    df_filtered["Name"] = df_filtered["Name"].astype(str).str.strip()
    return set(df_filtered["Name"].tolist())

correct_confidences = []
incorrect_confidences = []

cross_df = pd.read_excel("01_frog_data_compilation/results/cross_verification_results.xlsx")
ref_df = pd.read_excel("01_frog_data_compilation/data/Reference_Froggy_Spreadsheet.xlsx")
analysis_df = pd.read_excel("01_frog_data_compilation/results/froggy_analysis_results.xlsx")
confidence_df = pd.read_excel("01_frog_data_compilation/results/egg_style_confidence.xlsx")

ref_df["Name"] = ref_df["Name"].astype(str).str.strip()
ref_df["Egg Style"] = ref_df["Egg Style"].fillna("-").astype(str).str.strip()

analysis_df["Name"] = analysis_df["Name"].astype(str).str.strip()
analysis_df["Egg Style"] = analysis_df["Egg Style"].fillna("-").astype(str).str.strip()

cross_df["Name"] = cross_df["Name"].astype(str).str.strip()
if "Egg Style" not in cross_df.columns:
    cross_df["Egg Style"] = "-"

green_names = get_green_highlighted_names("01_frog_data_compilation/data/Reference_Froggy_Spreadsheet.xlsx")
analysis_df = analysis_df[analysis_df["Name"].isin(green_names)]

ref_egg_dict = ref_df.set_index("Name")["Egg Style"].to_dict()
analysis_egg_dict = analysis_df.set_index("Name")["Egg Style"].to_dict()
confidence_dict = confidence_df.set_index("Name")["Confidence"].to_dict()

def determine_egg_style(name):
    if name not in green_names:
        return "-"

    ref_val = ref_egg_dict.get(name, "-")
    analysis_val = analysis_egg_dict.get(name, "-")

    if ref_val == "-" or analysis_val == "-":
        return "-"

    try:
        ref_int = int(ref_val)
        analysis_int = int(analysis_val)
    except ValueError:
        return "?"

    confidence = confidence_dict.get(name, None)

    if (analysis_int == 0 and ref_int == 1) or (analysis_int == 1 and ref_int == 2):
        if confidence is not None and confidence != "-":
            correct_confidences.append(int(confidence))
        return str(analysis_int)
    else:
        if confidence is not None and confidence != "-":
            incorrect_confidences.append(int(confidence))
        return "invalid"

cross_df["Egg Style"] = cross_df["Name"].apply(determine_egg_style)
cross_df.to_excel("01_frog_data_compilation/results/cross_verification_results.xlsx", index=False)

print(correct_confidences)
print(sum(correct_confidences) / len(correct_confidences))

print(incorrect_confidences)
print(sum(incorrect_confidences) / len(incorrect_confidences))