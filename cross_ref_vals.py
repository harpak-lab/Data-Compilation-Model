import pandas as pd
from openpyxl import load_workbook
import os

def filter_big_spreadsheet(input_filepath, sheet_name, output_filepath):
    """Extracts only the green-highlighted rows from an Excel sheet and returns them as a DataFrame."""
    wb = load_workbook(filename=input_filepath, data_only=True)
    ws = wb[sheet_name]

    green_rows = [] # store highlighted entries
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.fill.start_color.index == "FF92D050": # hex for green
                green_rows.append(cell.row)
    
    df = pd.read_excel(input_filepath, sheet_name=sheet_name)
    
    # filter only green-highlighted rows
    df_filtered = df.iloc[[r - 2 for r in green_rows]]

    # uncertainty cols -- make blanks 0 for cross verification
    uncertainty_columns = ["+/- SVL Male (mm)", "+/- SVL Female (mm)", "+/- SVL Adult (mm)", "+/- Egg Diameter (mm)"]
    for col in uncertainty_columns:
        if col in df_filtered.columns:
            df_filtered[col] = df_filtered[col].fillna(0)

    df_filtered.to_excel(output_filepath, index=False)
    # print(f"Green-highlighted rows saved to {output_filepath}")

def filter_my_spreadsheet(my_spreadsheet, filtered_spreadsheet, output_filepath):
    df_my = pd.read_excel(my_spreadsheet)
    df_filtered = pd.read_excel(filtered_spreadsheet)
    
    # safety check: "name" column exists
    if "Name" not in df_my.columns or "Name" not in df_filtered.columns:
        raise ValueError("Missing 'Name' column in one of the spreadsheets.")
    
    # spreadsheet should only include species in filtered spreadsheet
    df_my_filtered = df_my[df_my["Name"].isin(df_filtered["Name"])]
    
    df_my_filtered.to_excel(output_filepath, index=False)
    # print(f"Filtered spreadsheet saved to {output_filepath}")

def compare_values(reference_spreadsheet, my_filtered_spreadsheet, output_filepath):
    # make dataframes
    df_ref = pd.read_excel(reference_spreadsheet)
    df_filtered = pd.read_excel(my_filtered_spreadsheet)
    
    # create comparison dataframe
    df_comparison = pd.DataFrame(columns=df_ref.columns)
    df_comparison['Name'] = df_ref['Name']
    
    # comparison type for column groups
    groups = [
        {
            'columns': ['SVL Male (mm)', '+/- SVL Male (mm)'],
            'type': 'value_uncertainty'
        },
        {
            'columns': ['SVL Female (mm)', '+/- SVL Female (mm)'],
            'type': 'value_uncertainty'
        },
        {
            'columns': ['Avg SVL Adult (mm)', '+/- SVL Adult (mm)'],
            'type': 'value_uncertainty'
        },
        {
            'columns': ['Avg Egg Diameter (mm)', '+/- Egg Diameter (mm)'],
            'type': 'value_uncertainty'
        },
        {
            'columns': ['Min Egg Clutch', 'Max Egg Clutch'],
            'type': 'min_max'
        }
    ]
    
    # check if two ranges overlap
    def ranges_overlap(a_low, a_high, b_low, b_high):
        return a_low <= b_high and b_low <= a_high
    
    # check if value is numeric
    def is_valid(value):
        if pd.isna(value):
            return False
        if isinstance(value, str) and value.strip() == '-':
            return False
        try:
            float(value)
            return True
        except:
            return False
    
    for i in range(len(df_ref)): # each row
        row_ref = df_ref.iloc[i]
        row_filtered = df_filtered.iloc[i]
        
        for group in groups:
            cols = group['columns']
            group_type = group['type']
            
            if group_type == 'value_uncertainty':
                ref_val = row_ref[cols[0]]
                ref_unc = row_ref[cols[1]]
                filtered_val = row_filtered[cols[0]]
                filtered_unc = row_filtered[cols[1]]
                
                # validity of all 4 vals
                valid = all(is_valid(value) for value in [ref_val, ref_unc, filtered_val, filtered_unc])
                
                if not valid:
                    df_comparison.at[i, cols[0]] = '-'
                    df_comparison.at[i, cols[1]] = '-'
                else:
                    # convert to floats
                    ref_val_f = float(ref_val)
                    ref_unc_f = float(ref_unc)
                    filtered_val_f = float(filtered_val)
                    filtered_unc_f = float(filtered_unc)
                    
                    # check if equal
                    if ref_val_f == filtered_val_f and ref_unc_f == filtered_unc_f:
                        df_comparison.at[i, cols[0]] = ref_val_f
                        df_comparison.at[i, cols[1]] = ref_unc_f
                    else:
                        # not equal -- overlap? calc ranges
                        ref_low = ref_val_f - ref_unc_f
                        ref_high = ref_val_f + ref_unc_f
                        filtered_low = filtered_val_f - filtered_unc_f
                        filtered_high = filtered_val_f + filtered_unc_f
                        
                        # determine overlap
                        overlap = ranges_overlap(ref_low, ref_high, filtered_low, filtered_high)
                        result = 'overlap' if overlap else 'invalid'
                        df_comparison.at[i, cols[0]] = result
                        df_comparison.at[i, cols[1]] = result
            
            elif group_type == 'min_max':
                # similar process to before
                ref_min = row_ref[cols[0]]
                ref_max = row_ref[cols[1]]
                filtered_min = row_filtered[cols[0]]
                filtered_max = row_filtered[cols[1]]
                
                valid = all(is_valid(value) for value in [ref_min, ref_max, filtered_min, filtered_max])
                
                if not valid:
                    df_comparison.at[i, cols[0]] = '-'
                    df_comparison.at[i, cols[1]] = '-'
                else:
                    ref_min_f = float(ref_min)
                    ref_max_f = float(ref_max)
                    filtered_min_f = float(filtered_min)
                    filtered_max_f = float(filtered_max)
                    
                    if ref_min_f == filtered_min_f and ref_max_f == filtered_max_f:
                        df_comparison.at[i, cols[0]] = ref_min_f
                        df_comparison.at[i, cols[1]] = ref_max_f
                    else:
                        overlap = ranges_overlap(ref_min_f, ref_max_f, filtered_min_f, filtered_max_f)
                        result = 'overlap' if overlap else 'invalid'
                        df_comparison.at[i, cols[0]] = result
                        df_comparison.at[i, cols[1]] = result
    
    df_comparison.to_excel(output_filepath, index=False)
    print(f"Comparative data saved to {output_filepath}")

big_spreadsheet = "Froggy_Spreadsheet.xlsx"  # reference spreadsheet
output_spreadsheet = "filtered_big_spreadsheet.xlsx"
my_spreadsheet = "egg_analysis_results.xlsx"
filtered_output = "filtered_results.xlsx"
comparison_output = "cross_verification_results.xlsx"

# full pipeline
filter_big_spreadsheet(big_spreadsheet, "All Frogs", output_spreadsheet)
filter_my_spreadsheet(my_spreadsheet, output_spreadsheet, filtered_output)
compare_values(output_spreadsheet, filtered_output, comparison_output)

# get rid of intermediary spreadsheets
if os.path.exists(output_spreadsheet):
    os.remove(output_spreadsheet)
if os.path.exists(filtered_output):
    os.remove(filtered_output)